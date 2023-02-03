import json
import uuid

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet


def create_xlsx_file():
    with open("../data/menu.json", mode="r") as f:
        content = f.read()

    id_ = uuid.uuid4()
    menus = json.loads(content)

    wb = Workbook()
    wb.remove(wb.active)

    wb.create_sheet("Меню")
    ws = wb.active
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 210
    ws.column_dimensions["F"].width = 10

    row = 0
    thin = Side(border_style="thin", color="00000000")
    double = Side(border_style="double", color="000000FF")
    border = Border(left=thin, right=thin, top=double, bottom=double)
    font = Font(bold=True)

    menu_fill = PatternFill("solid", fgColor="00FF9900")
    submenu_fill = PatternFill("solid", fgColor="0099CC00")
    dish_fill = PatternFill("solid", fgColor="00FFFF99")

    for menu_index, menu in enumerate(menus, start=1):
        row += 1
        construct_menu_cells(
            sheet=ws,
            index=menu_index,
            menu=menu,
            row=row,
            font=font,
            fill=menu_fill,
            border=border
        )

        for sub_index, submenu in enumerate(menu["submenus"], start=1):
            row += 1
            construct_submenu_cells(
                sheet=ws,
                index=sub_index,
                submenu=submenu,
                row=row,
                font=font,
                fill=submenu_fill,
                border=border
            )

            for dish_index, dish in enumerate(submenu["dishes"], start=1):
                row += 1
                construct_dish_cells(
                    sheet=ws,
                    index=dish_index,
                    dish=dish,
                    row=row,
                    font=font,
                    fill=dish_fill,
                    border=border
                )

    wb.save(f"../data/{id_}.xlsx")
    wb.close()


def construct_menu_cells(
        sheet: Worksheet,
        index: int,
        menu: dict,
        row: int,
        font: Font,
        fill: PatternFill,
        border: Border
) -> None:
    menu_cells = [
        sheet.cell(row, 1, index),
        sheet.cell(row, 2, menu["title"]),
        sheet.cell(row, 3, menu["description"])
    ]
    for cell in menu_cells:
        cell.font = font
        cell.fill = fill
        cell.border = border
    return None


def construct_submenu_cells(
        sheet: Worksheet,
        index: int,
        submenu: dict,
        row: int,
        font: Font,
        fill: PatternFill,
        border: Border
) -> None:
    submenu_cells = [
        sheet.cell(row, 2, index),
        sheet.cell(row, 3, submenu["title"]),
        sheet.cell(row, 4, submenu["description"])
    ]
    for cell in submenu_cells:
        cell.font = font
        cell.fill = fill
        cell.border = border
    return None


def construct_dish_cells(
        sheet: Worksheet,
        index: int,
        dish: dict,
        row: int,
        font: Font,
        fill: PatternFill,
        border: Border
) -> None:
    dish_cells = [
        sheet.cell(row, 3, index),
        sheet.cell(row, 4, dish["title"]),
        sheet.cell(row, 5, dish["description"]),
        sheet.cell(row, 6, dish["price"])
    ]
    for cell in dish_cells:
        cell.font = font
        cell.fill = fill
        cell.border = border
    return None


if __name__ == "__main__":
    create_xlsx_file()
